"""
按难度分析指标
"""

import openpyxl
import openpyxl.worksheet.worksheet
from config import AnalyseConfig
from sqlbox import SqlBox

workbook = openpyxl.load_workbook(AnalyseConfig.submitXlsx)

fullName = {"C1": None,
            "E1": None,
            "C2": None,
            "E2": None,
            "C3": None,
            "E3": None,
            "C4": None,
            "E4": None,
            "C5": None,
            "E5": None,
            "C6": None,
            "E6": None
            }


class Problem:
    match = None
    title = None

    def __init__(self, match, title):
        self.match = match
        self.title = title

probs = {
    "1": [
        Problem("C1", "A"), Problem("C1", "B"), Problem("C1", "C"),
        Problem("C1", "E"), Problem("C1", "F"), Problem("E1", "A"),
        Problem("E1", "B"), Problem("E1", "C"), Problem("E1", "D"),
        Problem("E1", "G"), Problem("C2", "A"), Problem("C2", "B"),
        Problem("E2", "A"), Problem("E2", "B"), Problem("E2", "C"),
        Problem("C3", "A"), Problem("C3", "C"), Problem("C3", "D"),
        Problem("C3", "F"), Problem("C3", "G"), Problem("E3", "A"),
        Problem("E3", "B"), Problem("E3", "C"), Problem("C4", "A"),
        Problem("C4", "C"), Problem("C4", "D"), Problem("E4", "C"),
        Problem("E4", "D"), Problem("C5", "A"), Problem("C5", "B"),
        Problem("C6", "A"), Problem("C6", "B"), Problem("C6", "D"),
        Problem("E6", "A"), Problem("E6", "B"), Problem("E6", "C")
    ],
    "2": [
        Problem("C1", "D"), Problem("C1", "G"), Problem("C1", "I"),
        Problem("E1", "E"), Problem("E1", "F"), Problem("E1", "H"),
        Problem("C2", "C"), Problem("C2", "D"), Problem("C2", "E"),
        Problem("C2", "F"), Problem("E2", "D"), Problem("E2", "E"),
        Problem("E2", "H"), Problem("C3", "B"), Problem("C3", "E"),
        Problem("E3", "D"), Problem("E3", "E"), Problem("C4", "B"),
        Problem("C4", "F"), Problem("E4", "E"), Problem("E4", "G"),
        Problem("C5", "C"), Problem("C5", "D"), Problem("C5", "F"),
        Problem("C5", "G"), Problem("E5", "A"), Problem("E5", "B"),
        Problem("C6", "C"), Problem("C6", "G")
    ],
    "3": [
        Problem("C1", "H"), Problem("C1", "J"), Problem("E1", "I"),
        Problem("E1", "J"), Problem("C2", "G"), Problem("C2", "H"),
        Problem("C2", "I"), Problem("E2", "F"), Problem("E2", "G"),
        Problem("C3", "H"), Problem("E3", "F"), Problem("E3", "I"),
        Problem("C4", "E"), Problem("C4", "G"), Problem("E4", "A"),
        Problem("E4", "B"), Problem("E4", "I"), Problem("E4", "J"),
        Problem("E5", "C"), Problem("E5", "D"), Problem("C6", "E"),
        Problem("C6", "F"), Problem("C6", "H"), Problem("C6", "I"),
        Problem("E6", "D"), Problem("E6", "E")
    ],
    "4": [
        Problem("C2", "J"), Problem("E2", "I"), Problem("E2", "J"),
        Problem("C3", "I"), Problem("C3", "J"), Problem("E3", "G"),
        Problem("E3", "H"), Problem("E3", "J"), Problem("E3", "K"),
        Problem("C4", "H"), Problem("C4", "I"), Problem("C4", "J"),
        Problem("E4", "F"), Problem("C5", "H"), Problem("C5", "J"),
        Problem("C5", "K"), Problem("E5", "E"), Problem("E5", "F"),
        Problem("E5", "G"), Problem("E5", "H"), Problem("C6", "J"),
        Problem("C6", "K"), Problem("E6", "F"), Problem("E6", "G"),
        Problem("E6", "H"), Problem("E6", "I"), Problem("E6", "J"),
    ],
    "5": [
        Problem("C3", "K"), Problem("E4", "K"), Problem("C5", "I"),
        Problem("E5", "I")
    ],
    "6": [
        Problem("E4", "H"), Problem("E5", "J")
    ],
    "7": [
        Problem("E6", "K")
    ]
}

staticSqlbox = SqlBox()
staticMatch = "xxx"
nowMatch = None

def calculate(match:str, title:str):
    sqlbox = fullName[match]

    problemsT = sqlbox.filter({"题目编号": title})
    submitSum = len(problemsT.value)
    studentSum = problemsT.Count("用户")

    problemsTAC = problemsT.filter({"结果": "Accepted"})
    studentSumAC = problemsTAC.Count("用户")

    return submitSum, studentSum, studentSumAC

if __name__ == '__main__':
    workbook = openpyxl.load_workbook(AnalyseConfig.submitXlsx)
    fullName["C1"] = SqlBox(workbook["C1-2021级-航类第1次上机"])
    print("load C1")
    fullName["E1"] = SqlBox(workbook["E1-2021级-航类第1次练习"])
    print("load E1")
    fullName["C2"] = SqlBox(workbook["C1-2021级-航类第1次上机"])
    print("load C2")
    fullName["E2"] = SqlBox(workbook["E2-2021级-航类第2次练习"])
    print("load E2")
    fullName["C3"] = SqlBox(workbook["C3-2021级-航类第3次上机"])
    print("load C3")
    fullName["E3"] = SqlBox(workbook["E3-2021级-航类第3次练习"])
    print("load E3")
    fullName["C4"] = SqlBox(workbook["C4-2021级-航类第4次上机"])
    print("load C4")
    fullName["E4"] = SqlBox(workbook["E4-2021级-航类第4次练习"])
    print("load E4")
    fullName["C5"] = SqlBox(workbook["C5-2021级-航类第5次上机"])
    print("load C5")
    fullName["E5"] = SqlBox(workbook["E5-2021级-航类第5次练习"])
    print("load E5")
    fullName["C6"] = SqlBox(workbook["C6-2021级-航类第6次上机"])
    print("load C6")
    fullName["E6"] = SqlBox(workbook["E6-2021级-航类第6次练习"])
    print("load E6")


    savebook = openpyxl.load_workbook(AnalyseConfig.hardXlsx)

    for key, value in probs.items():

        savesheet = savebook.create_sheet(key)
        col = ['比赛', '题号', '总提交数', '总提交人数', '通过人数', '通过率', '正确率']
        for k in range(0, len(col)):
            savesheet.cell(1, k + 1, col[k])

        proList = value

        i = 2
        for problem in proList:
            # worksheet = workbook[fullName[problem.match]]
            nowMatch = problem.match
            submitSum, studentSum, studentSumAC = calculate(problem.match, problem.title)

            print('{} {} {} {} {} {} {}'.format(problem.match, problem.title, submitSum, studentSum, studentSumAC, studentSumAC/studentSum, studentSumAC/submitSum))

            savesheet.cell(i, 1, problem.match)
            savesheet.cell(i, 2, problem.title)
            savesheet.cell(i, 3, submitSum)
            savesheet.cell(i, 4, studentSum)
            savesheet.cell(i, 5, studentSumAC)
            savesheet.cell(i, 6, studentSumAC/studentSum)
            savesheet.cell(i, 7, studentSumAC/submitSum)

            i = i + 1

    savebook.save(AnalyseConfig.hardXlsx)