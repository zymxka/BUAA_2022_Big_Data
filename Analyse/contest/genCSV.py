import sys

sys.path.append("../")

import openpyxl
from sqlbox import SqlBox as Sq
from config import AnalyseConfig
import matplotlib.pyplot as plt
import matplotlib
import seaborn as sns
import pandas as pd


def getGrade(path: str):
    stu = Sq(openpyxl.load_workbook(AnalyseConfig.rstuXlsx)["学生信息"]) \
        .filter({"身份": "member"}, 0) \
        .choose(["成员id"]).orderBy(["成员id"], reverse=True) \
        .Extend({"分数": 0})

    wb = openpyxl.load_workbook(path)
    for sheet in wb:
        pb = Sq(sheet).choose(["用户id", "分数"]).toFloat("分数") \
            .LinkWith(stu, "成员id", "用户id").ModifyHead({"left.用户id": "用户id", "right.分数": "分数"}) \
            .orderBy(["用户id", "分数"], reverse=True).unique("用户id")
        i = 0
        for j in pb.value:
            while i < len(stu.value) and stu.value[i]["成员id"] != j["用户id"]:
                i += 1
            if i >= len(stu.value):
                break
            stu.value[i]["分数"] += j["分数"]

    stu = stu.orderBy("分数", reverse=True)
    result = []
    for i in stu:
        result.append(i["分数"])
    return result


def data(idL=0, idR=22000000, title=''):
    stu = Sq(openpyxl.load_workbook(AnalyseConfig.rstuXlsx)["学生信息"]) \
        .filter({"身份": "member"}, 0) \
        .filter({"学号": idL}, 1).filter({"学号": idR}, -2) \
        .choose(["成员id"]).orderBy(["成员id"], reverse=True).Extend({"分数": 0})

    wb = openpyxl.load_workbook("../../Table/rank.xlsx")
    dataC = {"contest": [], "grade": []}
    dataE = {"contest": [], "grade": []}
    dataCE = {"contest": [], "grade": []}
    for sheet in wb:
        name = sheet.title
        rnk = Sq(sheet).choose(["学号", "得分"]).LinkWith(stu, "学号", "成员id")
        if not name.endswith('上机'):
            for i in rnk.value:
                dataE["grade"].append(float(i["left.得分"]))
                dataE["contest"].append(name[:2])
        if not name.endswith('练习'):
            for i in rnk.value:
                dataC["grade"].append(float(i["left.得分"]))
                dataC["contest"].append(name[:2])
        for i in rnk.value:
            dataCE["grade"].append(float(i["left.得分"]))
            dataCE["contest"].append(name[:2])
    ds = pd.DataFrame(dataC)
    ds.to_csv(f"dataC{title}.csv", index=False)
    ds = pd.DataFrame(dataE)
    ds.to_csv(f"dataE{title}.csv", index=False)
    ds = pd.DataFrame(dataCE)
    ds.to_csv(f"dataCE{title}.csv", index=False)
    # pd.DataFrame({"lab": labels}).to_csv("boxplot_lab.csv", index=False)


if __name__ == "__main__":
    # data(21374000, 21375000, '2174')
    # data(21375000, 21376000, '2175')
    # data(21376000, 21377000, '2176')
    # data(21180000, 21190000, '2118')
    # data(21375410, 21375439, '2175xwc')
    data(21040000, 21050000, '2104xwc')

