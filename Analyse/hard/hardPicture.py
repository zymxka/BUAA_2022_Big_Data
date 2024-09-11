import locale

import openpyxl
from mpl_toolkits.mplot3d import Axes3D

from sqlbox import SqlBox
from config import AnalyseConfig
import matplotlib.pyplot as plt
import numpy as np
from scipy.interpolate import griddata

plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
plt.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号
color = ['#F1948A', '#C39BD3', '#AED6F1', '#A3E4D7', '#F9E79F', '#EDBB99', '#ABB2B9']


def passTimeChange2():
    workbook = openpyxl.load_workbook(AnalyseConfig.hardXlsx)
    index1 = 0
    for sheet in workbook:
        match = []
        number = []
        rate = []
        sqlbox = SqlBox(sheet)
        bk = sqlbox.value
        for value in bk:
            if value['比赛'] in match:
                index = 0
                for a in match:
                    if a == value['比赛']:
                        break
                    else:
                        index += 1
                rate[index] = rate[index] + value['通过率']
                number[index] = number[index] + 1
            else:
                match.append(value['比赛'])
                rate.append(value['通过率'])
                number.append(1)
        length = len(match)
        c = []
        cr = []
        e = []
        er = []
        for i in range(length):
            rate[i] = rate[i] / float(number[i])
        for i in range(length):
            if match[i][0] == 'C':
                c.append(match[i])
                cr.append(rate[i])
            else:
                e.append(match[i])
                er.append(rate[i])
        plt.ylim(0, 1)
        plt.plot(c, cr, color=color[index1])
        plt.title("hard:" + str(index1 + 1) + " C")
        plt.savefig("../Output/hard/pass/" + str(index1 + 1) + "C")
        plt.close()
        plt.ylim(0, 1)
        plt.plot(e, er, color=color[index1])
        plt.title("hard:" + str(index1 + 1) + " E")
        plt.savefig("../Output/hard/pass/" + str(index1 + 1) + "E")
        plt.close()
        index1 += 1


def accuracyChange2():
    workbook = openpyxl.load_workbook(AnalyseConfig.hardXlsx)
    color = ['#F1948A', '#C39BD3', '#AED6F1', '#A3E4D7', '#F9E79F', '#EDBB99', '#ABB2B9']
    index1 = 0
    for sheet in workbook:
        match = []
        number = []
        rate = []
        sqlbox = SqlBox(sheet)
        bk = sqlbox.value
        for value in bk:
            if value['比赛'] in match:
                index = 0
                for a in match:
                    if a == value['比赛']:
                        break
                    else:
                        index += 1
                rate[index] = rate[index] + value['正确率']
                number[index] = number[index] + 1
            else:
                match.append(value['比赛'])
                rate.append(value['正确率'])
                number.append(1)
        length = len(match)
        c = []
        cr = []
        e = []
        er = []
        for i in range(length):
            rate[i] = rate[i] / float(number[i])
        for i in range(length):
            if match[i][0] == 'C':
                c.append(match[i])
                cr.append(rate[i])
            else:
                e.append(match[i])
                er.append(rate[i])
        plt.ylim(0, 1)
        plt.plot(c, cr, color=color[index1])
        plt.title("hard:" + str(index1 + 1) + " C")
        plt.savefig("../Output/hard/accuracy/" + str(index1 + 1) + "C")
        plt.close()
        plt.ylim(0, 1)
        plt.plot(e, er, color=color[index1])
        plt.title("hard:" + str(index1 + 1) + " E")
        plt.savefig("../Output/hard/accuracy/" + str(index1 + 1) + "E")
        plt.close()
        index1 += 1


def passTimeChange3():
    workbook = openpyxl.load_workbook(AnalyseConfig.hardXlsx)
    color = ['r', 'g', 'b', '#A3E4D7', '#F9E79F', '#EDBB99', '#ABB2B9']
    index1 = 0
    index2 = 0
    fig = plt.figure()
    ax = fig.add_subplot(111, projection='3d')
    rate = []
    order = []
    hard = []
    for sheet in workbook:
        index1 += 1
        sqlbox = SqlBox(sheet)
        bk = sqlbox.value
        for value in bk:
            rate.append(value['通过率'])
            order.append(ord(value['题号']) - 64)
            hard.append(index1)
    ax.plot_trisurf(order, hard, rate, color='#A3E4D7')
    ax.set_xlabel('order', color='#D7BDE2')
    ax.set_ylabel('hard', color='#EDBB99')
    ax.set_zlabel('rate', color='#85C1E9')
    plt.draw()
    plt.show()


def accuracyChange3():
    workbook = openpyxl.load_workbook(AnalyseConfig.hardXlsx)
    color = ['r', 'g', 'b', '#A3E4D7', '#F9E79F', '#EDBB99', '#ABB2B9']
    index1 = 0
    index2 = 0
    fig = plt.figure()
    ax = fig.add_subplot(111, projection='3d')
    rate = []
    order = []
    hard = []
    for sheet in workbook:
        index1 += 1
        sqlbox = SqlBox(sheet)
        bk = sqlbox.value
        for value in bk:
            rate.append(value['正确率'])
            order.append(ord(value['题号']) - 64)
            hard.append(index1)
    ax.plot_trisurf(order, hard, rate, color='#F9E79F')
    ax.set_xlabel('order', color='#D7BDE2')
    ax.set_ylabel('hard', color='#EDBB99')
    ax.set_zlabel('rate', color='#85C1E9')
    plt.draw()
    plt.show()


def hard_distribution():
    hard = [1, 2, 3, 4, 5, 6, 7]
    num = [36, 29, 26, 27, 4, 2, 1]
    plt.bar(hard, num, width=0.7, color=color[0])
    plt.xlabel("题目难度")
    plt.ylabel("数量")
    for index, value in enumerate(num):
        plt.text(index + 1, value + 0.5,
                 str(value))
    plt.savefig("所有题目的难度分布")
    plt.show()


def hard_number():
    workbook = openpyxl.load_workbook(AnalyseConfig.hardXlsx)
    index1 = 0
    hard = []
    c_p = []
    e_p = []
    for sheet in workbook:
        match = []
        number = []
        rate = []
        sqlbox = SqlBox(sheet)
        bk = sqlbox.value
        for value in bk:
            if value['比赛'] in match:
                index = 0
                for a in match:
                    if a == value['比赛']:
                        break
                    else:
                        index += 1
                rate[index] = rate[index] + value['通过人数']
                number[index] = number[index] + 1
            else:
                match.append(value['比赛'])
                rate.append(value['通过人数'])
                number.append(1)
        length = len(match)
        c = 0
        cr = 0
        e = 0
        er = 0
        for i in range(length):
            if match[i][0] == 'C':
                c = c + number[i]
                cr = cr + rate[i]
            else:
                e = e + number[i]
                er = er + rate[i]
        hard.append(index1 + 1)
        if c != 0:
            c_p.append(cr / c)
        else:
            c_p.append(0)
        if e != 0:
            e_p.append(er / e)
        else:
            e_p.append(0)
        index1 += 1
    plt.plot(hard, c_p, color='#F1948A', label="C")
    plt.plot(hard, e_p, color='#C39BD3', label="E")
    plt.xlabel("难度")
    plt.ylabel("通过人数")
    plt.legend()
    for i in range(0, 7):
        plt.text(hard[i], int(c_p[i]), str(int(c_p[i])))
    for i in range(0, 7):
        plt.text(hard[i], int(e_p[i])+2, str(int(e_p[i])))
    plt.savefig("难度与通过人数")
    plt.show()


if __name__ == '__main__':
    hard_number()
