import sys

sys.path.append("../")

import openpyxl
from sqlbox import SqlBox
from config import AnalyseConfig


def getid():
    workbook2 = openpyxl.load_workbook(AnalyseConfig.rstuXlsx)
    sheet0 = workbook2["学生信息"]
    sqlboxstu = SqlBox(sheet0)

    stu = sqlboxstu.filter({"身份": "member"}, 0).choose(["成员id"]).value
    savebook = openpyxl.load_workbook("stuid.xlsx")
    savesheet = savebook.create_sheet("排名信息")
    col = ["学生id"]
    for k in range(0, len(col)):
        savesheet.cell(1, k + 1, col[k])
    for idx, i in enumerate(stu):
        col = [i['成员id']]
        for k in range(0, len(col)):
            savesheet.cell(idx + 2, k + 1, col[k])
    savebook.save("stuid.xlsx")


def getGrade():
    workbook1 = openpyxl.load_workbook(AnalyseConfig.rranXlsx)
    workbook2 = openpyxl.load_workbook("stuid.xlsx")
    sqlboxstu = SqlBox(workbook2['排名信息'])

    ans = []
    for i in sqlboxstu.value:
        ans.append([i['学生id']])

    for sheet in workbook1:
        sqlboxrk = SqlBox(sheet)
        bk = sqlboxstu.LinkWith(sqlboxrk, '学生id', '学号').choose(["left.学生id", "right.得分"]).value
        print(len(bk))
        idx = 0
        seen = set()
        for i in bk:
            if i['left.学生id'] in seen:
                continue
            while i['left.学生id'] != ans[idx][0]:
                ans[idx].append(0)
                seen.add(ans[idx][0])
                idx += 1
            ans[idx].append(i['right.得分'])
            seen.add(ans[idx][0])
            idx += 1

    savebook = openpyxl.load_workbook("stugrade.xlsx")
    savesheet = savebook.create_sheet("得分")
    col = ["学生id", "C1", "E1", "C2", "E2", "C3", "E3", "C4", "E4", "C5", "E5", "Mid", "C6", "E6"]
    for k in range(0, len(col)):
        savesheet.cell(1, k + 1, col[k])
    for idx, i in enumerate(ans):
        col = i
        for k in range(0, len(col)):
            savesheet.cell(idx + 2, k + 1, col[k])
    savebook.save("stugrade.xlsx")


def getRank():
    workbook1 = openpyxl.load_workbook(AnalyseConfig.rranXlsx)
    workbook2 = openpyxl.load_workbook("stuid.xlsx")
    sqlboxstu = SqlBox(workbook2['排名信息'])

    ans = []
    for i in sqlboxstu.value:
        ans.append([i['学生id']])

    for sheet in workbook1:
        sqlboxrk = SqlBox(sheet)
        bk = sqlboxstu.LinkWith(sqlboxrk, '学生id', '学号').choose(["left.学生id", "right.排名"]).value
        print(len(bk))
        idx = 0
        seen = set()
        for i in bk:
            if i['left.学生id'] in seen:
                continue
            while i['left.学生id'] != ans[idx][0]:
                ans[idx].append(9999)
                seen.add(ans[idx][0])
                idx += 1
            ans[idx].append(i['right.排名'])
            seen.add(ans[idx][0])
            idx += 1

    savebook = openpyxl.load_workbook("sturk.xlsx")
    savesheet = savebook.create_sheet("排名")
    col = ["学生id", "C1", "E1", "C2", "E2", "C3", "E3", "C4", "E4", "C5", "E5", "Mid", "C6", "E6"]
    for k in range(0, len(col)):
        savesheet.cell(1, k + 1, col[k])
    for idx, i in enumerate(ans):
        col = i
        for k in range(0, len(col)):
            savesheet.cell(idx + 2, k + 1, col[k])
    savebook.save("sturk.xlsx")


def standard_grade():
    book = openpyxl.load_workbook("stugrade.xlsx")
    sheet = book['得分']
    savesheet = book.create_sheet("归一化得分")

    data = []

    for row in sheet.rows:
        tmp = []
        for item in row:
            tmp.append(item.value)
        data.append(tmp)

    for col in range(1, len(data[0])):
        tmp = []
        for row in range(1, len(data)):
            tmp.append(float(data[row][col]))
        tmp_min = min(tmp)
        tmp_max = max(tmp)
        for i in range(len(tmp)):
            tmp[i] = (tmp[i] - tmp_min) / (tmp_max - tmp_min)

        for row in range(1, len(data)):
            data[row][col] = tmp[row - 1]

    #print(data)

    for i in range(len(data)):
        for j in range(len(data[0])):
            savesheet.cell(i + 1, j + 1, data[i][j])
    book.save("stugrade.xlsx")



if __name__ == "__main__":
    # getid()
    standard_grade()
    getid()
