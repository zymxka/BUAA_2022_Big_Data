import functools

import openpyxl.worksheet.worksheet
import openpyxl


class SqlBox:
    """
    模拟SQL操作
    """

    value = None

    def __init__(self, sheet: openpyxl.worksheet.worksheet.Worksheet=None):
        if sheet != None:
            columns = []
            for i in range(1, sheet.max_column + 1):
                columns.append(sheet.cell(1, i).value)

            results = []
            for i in range(2, sheet.max_row + 1):
                result = {}
                for j in range(0, sheet.max_column):
                    value = sheet.cell(i, j + 1).value
                    result[columns[j]] = value
                results.append(result)

            self.value = results
        else: self.value = []

    # 判断一个字符串是否是数字
    def is_number(self,s):
        try:
            float(s)
            return True
        except ValueError:
            pass
        return False


    # select {columns} from
    def choose(self, columns:list):
        results = []
        for item in self.value:
            result = {}
            for column in columns:
                result[column] = item[column]
            results.append(result)

        sqlbox = SqlBox()
        sqlbox.value = results
        return sqlbox

    # where {filters}
    # filters: {'column1': 'value1', 'colomn2': 'value2', ...}
    # equal: {-2:<; -1:<=; 0:=; 1:>=; 2:>; 3:!=}
    def filter(self, filters:dict, equal:int=0):
        results = []
        for item in self.value:
            choose = True
            for key, value in filters.items():
                if self.is_number(value):
                    if (float(item[key]) >= float(value) and equal == -2)\
                            or (float(item[key]) > float(value) and equal == -1)\
                            or (item[key] != value and equal == 0)\
                            or (float(item[key]) < float(value) and equal == 1)\
                            or (float(item[key]) <= float(value) and equal == 2)\
                            or (item[key] == value and equal == 3):
                        choose = False
                        break
                else:
                    if item[key] != value and equal == 0:
                        choose = False
            if choose:
                results.append(item)

        sqlbox = SqlBox()
        sqlbox.value = results
        return sqlbox

    # order by {keys}
    # keys: ['column1', 'column2', ...]
    # reverse: {False: asc; True: desc}
    def orderBy(self, keys: list, reverse: bool=False):
        def cmp(item1: dict, item2: dict):
            for key in keys:
                if float(item1[key]) < float(item2[key]):
                    return -1
                elif float(item1[key]) > float(item2[key]):
                    return 1
            return 0

        sqlbox = SqlBox()
        sqlbox.value = sorted(self.value, key=functools.cmp_to_key(cmp), reverse=reverse)
        return sqlbox

    def unique(self, *keys):
        prv = {}
        for k in keys:
            prv[k] = None

        def cmp(x):
            ans = False
            for k in keys:
                if prv[k] != x[k]:
                    ans = True
                    prv[k] = x[k]
            return ans
        result = SqlBox()
        for i in self.value:
            if cmp(i):
                result.value.append(i)
        return result

    # count(column)
    def Count(self, column: str):
        count = set()
        for item in self.value:
            if item[column] not in count:
                count.add(item[column])
        return len(count)

    # sum(column)
    def Sum(self, column: str):
        sum = 0
        for item in self.value:
            sum = sum + float(item[column])
        return sum

    # avg(column)
    def Avg(self, column: str):
        return self.Sum(column) / len(self.value)

    # max(column)
    def Max(self, column: str):
        max = -float('inf')
        for item in self.value:
            if float(item[column]) > max:
                max = float(item[column])
        return max

    # min(column)
    def Min(self, column: str):
        min = float('inf')
        for item in self.value:
            if float(item[column]) < min:
                min = float(item[column])
        return min

    def Extend(self, kv: dict):
        result = SqlBox()
        for i in self.value:
            j = {}
            for k, v in i.items():
                j[k] = v
            for k, v in kv.items():
                j[k] = v
            result.value.append(j)
        return result

    # select * from sheet1, sheet2 where sheet1.column1 = sheet2.column2
    def LinkWith(self, sqlbox, column1: str, column2: str):
        results = []
        for item1 in self.value:
            for item2 in sqlbox.value:
                if item1[column1] == item2[column2]:
                    result = {}
                    for key, value in item1.items():
                        result["left." + key] = value
                    for key, value in item2.items():
                        result["right." + key] = value
                    results.append(result)

        sqlbox0 = SqlBox()
        sqlbox0.value = results
        return sqlbox0

    # select a as b from sheet
    def ModifyHead(self, modify: dict):
        results = []
        for item in self.value:
            result = {}
            for key, value in item.items():
                if key in modify.keys():
                    result[modify[key]] = value
                else:
                    result[key] = value
            results.append(result)

        sqlbox = SqlBox()
        sqlbox.value = results
        return sqlbox

    # union
    def Union(self, sqlbox):
        results = list(set(self.value).union(sqlbox.value))

        sqlbox0 = SqlBox()
        sqlbox0.value = results
        return sqlbox0

    # intersect
    def Intersect(self, sqlbox):
        results = list(set(self.value).intersection(sqlbox.value))

        sqlbox0 = SqlBox()
        sqlbox0.value = results
        return sqlbox0

    # diff
    def Diff(self, sqlbox):
        results = list(set(self.value).difference(sqlbox.value))

        sqlbox0 = SqlBox()
        sqlbox0.value = results
        return sqlbox0

    def toFloat(self, *keys):
        for i in self.value:
            for k in keys:
                i[k] = float(i[k])
        return self
