import sys
import time
import openpyxl.worksheet.worksheet
import openpyxl
import pickle

sys.path.append("../")

from sqlbox import SqlBox as Sq
from config import AnalyseConfig


class SLog:
    INF = time.mktime(time.strptime("2099-01-01", '%Y-%m-%d'))

    def __init__(self, uid):
        self.uid = int(uid)
        self.fr, self.to = self.INF, self.INF
        self.cnt, self.ac = 0, False

    def rst(self):
        self.fr, self.to = self.INF, self.INF
        self.cnt, self.ac = 0, False

    @staticmethod
    def time(t: str):
        return time.mktime(time.strptime(t, '%Y-%m-%d %X'))

    def upd(self, stat: str, t: str):
        self.cnt += int(not self.ac)
        self.fr = min(self.fr, self.time(t))
        if stat == 'Accepted':
            self.to = min(self.to, self.time(t))
            self.ac = True
        return self

    def status(self):
        delta, st = self.to - self.fr, True
        if not self.to < self.INF:
            delta, st = self.INF, False
        return dict(id=self.uid, ac=st, cnt=self.cnt, dt=delta)


class Calc:
    def __init__(self, idL=0, idR=22000000, cls: str = ''):
        self.cls = cls
        sheet = openpyxl.load_workbook(AnalyseConfig.rstuXlsx)["学生信息"]
        self.map, self.log = dict(), []
        self.idL, self.idR = idL, idR
        for i in range(2, sheet.max_row + 1):
            self.map[sheet.cell(i, 1).value] = i - 2
            self.log.append(SLog(sheet.cell(i, 3).value))

    def rst(self):
        for i in self.log:
            i.rst()

    def run(self, contest: str, path: str):
        wb = openpyxl.load_workbook(path)
        # data = dict(pid=[], result=[])
        data = dict(pid=[], tot=[], ac=[], cnta=[], cntw=[])
        for sheet in wb:
            self.rst()
            pid = sheet.title
            lst = Sq(sheet)
            for i in lst.value:
                i['t'] = SLog.time(i['提交时间'])
            lst.value.sort(key=lambda x: x['t'])
            for i in lst.value:
                self.log[self.map[i['用户id']]].upd(i['结果'], i['提交时间'])
            data['pid'].append(pid)
            st = self.stat()
            for k, v in st.items():
                data[k].append(v)
            # result = Sq()
            # for i in self.log:
            #     result.value.append(i.status())
            # data['pid'].append(pid)
            # data['sq'].append(result)
            # print(pid, self.log[self.map['21375424']].status())
        print(data)
        pickle.dump(data, open(f"{contest}{self.cls}_ac.pkl", "wb"), -1)

    def stat(self):
        ac, cnta, cntw, siz = 0, 0, 0, 0
        for i in self.log:
            if self.idL <= i.uid < self.idR:
                st = i.status()
                ac += int(st['ac'])
                siz += int(st['cnt'] != 0)
                cnta += st['cnt'] if st['ac'] else 0
                cntw += 0 if st['ac'] else st['cnt']
        return dict(tot=siz, ac=ac, cnta=cnta, cntw=cntw)


cal = Calc(21180000, 21190000, '2118')
for c in ['C1', 'C2', 'C3', 'C4', 'C5', 'C6', 'E1', 'E2', 'E3', 'E4', 'E5', 'E6']:
    cal.run(c, f"../../Table/单个题目的提交信息/{c}.xlsx", )
