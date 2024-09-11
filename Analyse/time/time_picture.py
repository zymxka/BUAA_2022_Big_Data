# -*- coding: utf-8 -*-

import openpyxl
import matplotlib.pyplot as plt

plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
plt.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号
time = ["0-5", "5-10", "10-15", "15-20", "20-25", "25-30", "30-35", "35-40", "40-45", "45-50", "50-55", "55-60",
        "60-65", "65-70", "70-75", "75-80", "80-85", "85-90", "90-95", "95-100", "100-105", "105-110", "110-115",
        "115-120"]

color = ['#F1948A', '#C39BD3', '#AED6F1', '#A3E4D7', '#F9E79F', '#EDBB99', '#ABB2B9', '#008B8B', "#9ACD32", '#FFD700',
         '#F4A460']


def has_k():
    sheet = ["C3", "C5", "C6"]
    workbook = openpyxl.load_workbook("../../Table/timer.xlsx")

    label = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
    for i in sheet:
        index = 0
        name = i
        for row in workbook[i].iter_rows(min_row=2, max_row=12, min_col=2, max_col=25):
            num = []
            for i in row:
                num.append(i.value)
            plt.plot(time, num, color=color[index], label=label[index])
            index += 1
        plt.xlabel("时间段")
        plt.ylabel("通过人数")
        plt.legend()
        plt.title("竞赛时间与通过人数: " + name)
        plt.show()
        plt.close()


def no_k():
    sheet = ["C1", "C2", "C4"]
    workbook = openpyxl.load_workbook("../../Table/timer.xlsx")
    label = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]
    for i in sheet:
        name = i
        index = 0
        for row in workbook[i].iter_rows(min_row=2, max_row=11, min_col=2, max_col=25):
            num = []
            for i in row:
                num.append(i.value)
            plt.plot(time, num, color=color[index], label=label[index])
            index += 1
        plt.xlabel("时间段")
        plt.ylabel("通过人数")
        plt.legend()
        plt.title("竞赛时间与通过人数: " + name)
        plt.show()
        plt.close()


def mid():
    sheet = ["Mid"]
    workbook = openpyxl.load_workbook("../../Table/timer.xlsx")
    label = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
    for i in sheet:
        index = 0
        name = i
        for row in workbook[i].iter_rows(min_row=2, max_row=12, min_col=2, max_col=25):
            num = []
            for i in row:
                num.append(i.value)
            plt.plot(time, num, color=color[index], label=label[index])
            index += 1
        plt.xlabel("时间段")
        plt.ylabel("通过人数")
        plt.legend()
        plt.title("竞赛时间与通过人数: " + name)
        plt.show()
        plt.close()


if __name__ == "__main__":
    mid()
