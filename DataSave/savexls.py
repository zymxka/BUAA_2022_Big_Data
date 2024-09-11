import xlrd
import xlwt
import openpyxl
from DataSave.config import DataSaveConfig
from Entity.SubmitItem import SubmitItem
from Entity.SubmitTitle import SubmitTitle
from Entity.StudentTitle import StudentTitle


def saveRankXls(rankTitle, rankInfos):
    workbook = openpyxl.load_workbook(DataSaveConfig.rankXlsx)
    worksheet = workbook.create_sheet(rankTitle.contName)
    col = []
    col.append(rankTitle.rankTitle)
    col.append(rankTitle.studentIdTitle)
    col.append(rankTitle.scoreTitle)
    col.append(rankTitle.penaltyTitle)
    for result in rankTitle.resultTitles:
        col.append(result)

    for k in range(0, len(col)):
        worksheet.cell(1, k + 1, col[k])

    length = len(rankInfos)
    for i in range(0, length):
        rankItem = rankInfos[i]
        col = []
        col.append(rankItem.rank)
        col.append(rankItem.studentId)
        col.append(rankItem.score)
        col.append(rankItem.penalty)
        for result in rankItem.results:
            col.append(result)
        for k in range(0, len(col)):
            worksheet.cell(i + 2, k + 1, col[k])
    workbook.save(DataSaveConfig.rankXlsx)


def saveSubmitXls(submitInfos, contName):
    workbook = openpyxl.load_workbook(DataSaveConfig.submitXlsx)
    worksheet = workbook.create_sheet(contName)
    col = []
    col.append(SubmitTitle.probIdTitle)
    col.append(SubmitTitle.userNameTitle)
    col.append(SubmitTitle.resultTitle)
    col.append(SubmitTitle.scoreTitle)
    col.append(SubmitTitle.codeLenTitle)
    col.append(SubmitTitle.runTimeTitle)
    col.append(SubmitTitle.memoryTitle)
    col.append(SubmitTitle.timeTitle)

    for k in range(0, len(col)):
        worksheet.cell(1, k + 1, col[k])

    length = len(submitInfos)
    for i in range(0, length):
        submitItem = submitInfos[i]
        col = []
        col.append(submitItem.probId)
        col.append(submitItem.userName)
        col.append(submitItem.result)
        col.append(submitItem.score)
        col.append(submitItem.codeLen)
        col.append(submitItem.runTime)
        col.append(submitItem.memory)
        col.append(submitItem.time)

        for k in range(0, len(col)):
            worksheet.cell(i + 2, k + 1, col[k])
    workbook.save(DataSaveConfig.submitXlsx)


def SaveProinfo(proinfo):
    workbook = openpyxl.load_workbook(DataSaveConfig.proXlsx)
    worksheet = workbook.create_sheet("题目提交信息")
    col = ['题目id', '题目名称', '作者id', '通过人数', '提交人数', '提交次数']
    for k in range(0, len(col)):
        worksheet.cell(1, k + 1, col[k])

    for idx, i in enumerate(proinfo):
        col = [i.proid, i.title, i.author, i.peoplepass, i.peoplesum, i.submitsum]
        for k in range(0, len(col)):
            worksheet.cell(idx + 2, k + 1, col[k])
    workbook.save(DataSaveConfig.proXlsx)


def SaveProsubmit(submits, title, path):
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.create_sheet(title)
    col = ['提交id', '用户id', '结果', '分数', '代码长度', '运行时间', '记忆', '提交时间']
    for k in range(0, len(col)):
        worksheet.cell(1, k + 1, col[k])

    for idx, i in enumerate(submits):
        col = [i.probId, i.userName, i.result, i.score, i.codeLen, i.runTime, i.memory, i.time]
        for k in range(0, len(col)):
            worksheet.cell(idx + 2, k + 1, col[k])

    workbook.save(path)


def SaveProdetail(info,path,title):
    workbook = openpyxl.load_workbook(path)
    worksheet = workbook.create_sheet(title)
    col = ['题目id', '难度', '标签1', '标签2', '标签3']
    for k in range(0, len(col)):
        worksheet.cell(1, k + 1, col[k])
        for idx, i in enumerate(info):
            for k in range(0, len(i)):
                worksheet.cell(idx + 2, k + 1, i[k])
    workbook.save(path)


def saveStudentXls(studentInfos):
    workbook = openpyxl.load_workbook(DataSaveConfig.studentXlsx)
    worksheet = workbook.create_sheet("学生信息")
    col = [StudentTitle.idTitle, StudentTitle.nameTitle, StudentTitle.sidTitle, StudentTitle.identityTitle]

    for k in range(0, len(col)):
        worksheet.cell(1, k + 1, col[k])

    length = len(studentInfos)
    for i in range(0, length):
        studentInfo = studentInfos[i]
        col = [studentInfo.id, studentInfo.name, studentInfo.sid, studentInfo.identity]

        for k in range(0, len(col)):
            worksheet.cell(i + 2, k + 1, col[k])
    workbook.save(DataSaveConfig.studentXlsx)
